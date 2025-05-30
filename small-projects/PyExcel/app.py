import openpyxl

wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]

# wb.create_chartsheet("Sheet2", 0)
# wb.remove("Sheet2")
cell = sheet["a1"]
print(cell.value)
print(cell.row)
print(cell.column)
print(cell.coordinate)

cell = sheet.cell(row=1, column=1)
print(sheet.max_row)
print(sheet.max_column)

# for row in range(1, sheet.max_row+1):
#     values = []
#     for column in range(1, sheet.max_column+1):
#         values.append(sheet.cell(row=row, column=column).value)
#         # print(sheet.cell(row=row, column=column).value)
#     print(values)

cell = sheet["a1"]
print(cell)

column = sheet["a"]
print("column: ", column)

cells = sheet["a:c"]
print("cells: ", cells)

print(sheet["a1:c3"])
print(sheet[1:3])

sheet.append([1, 2, 3])
wb.save("transactions2.xlsx")
